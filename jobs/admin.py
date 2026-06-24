from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import Application, CompanyLead


def export_to_excel(modeladmin, request, queryset):
    """
    Acción personalizada para exportar registros a Excel con hipervínculos a CVs
    """
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"
    
    # Obtener los nombres de los campos a exportar
    fields = [field.name for field in modeladmin.model._meta.fields]
    
    # Crear encabezados con estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, field_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = field_name.upper()
        cell.fill = header_fill
        cell.font = header_font
    
    # Agregar datos
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = getattr(obj, field_name)
            
            # Si el campo es un archivo CV y el objeto tiene ID, crear hipervínculo
            if field_name == 'cv' and value and hasattr(obj, 'id'):
                cv_url = request.build_absolute_uri(
                    reverse('jobs:download_cv', args=[obj.id])
                )
                cell.value = f"Descargar CV"
                cell.hyperlink = cv_url
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.value = str(value) if value else ""
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=registros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb.save(response)
    return response


export_to_excel.short_description = "Descargar registros en Excel"





class DateRangeFilter(admin.SimpleListFilter):
    parameter_name = 'created_at_range'
    template = 'admin/date_range_filter.html'

    def __init__(self, request, params, model, model_admin):
        self.request = request
        
        # Helper to pop values safely (handling strings, QueryDict, lists)
        def get_val(key):
            val = params.pop(key, None)
            if isinstance(val, list):
                return val[-1] if val else None
            return val

        self.range_val = get_val('created_at_range')
        self.gte_val = get_val('created_at__gte')
        self.lte_val = get_val('created_at__lte')
        
        super().__init__(request, params, model, model_admin)
        if self.range_val:
            self.used_parameters['created_at_range'] = self.range_val
        if self.gte_val:
            self.used_parameters['created_at__gte'] = self.gte_val
        if self.lte_val:
            self.used_parameters['created_at__lte'] = self.lte_val

    def lookups(self, request, model_admin):
        return ()

    def has_output(self):
        return True

    def queryset(self, request, queryset):
        from django.utils import timezone
        import datetime
        
        today = timezone.localdate()
        
        if self.range_val == 'today':
            queryset = queryset.filter(created_at__date=today)
        elif self.range_val == 'yesterday':
            yesterday = today - datetime.timedelta(days=1)
            queryset = queryset.filter(created_at__date=yesterday)
        elif self.range_val == '7_days':
            seven_days_ago = today - datetime.timedelta(days=7)
            queryset = queryset.filter(created_at__date__gte=seven_days_ago, created_at__date__lte=today)
        elif self.range_val == '30_days':
            thirty_days_ago = today - datetime.timedelta(days=30)
            queryset = queryset.filter(created_at__date__gte=thirty_days_ago, created_at__date__lte=today)
        elif self.range_val == 'this_month':
            queryset = queryset.filter(created_at__year=today.year, created_at__month=today.month)
        elif self.range_val == 'this_year':
            queryset = queryset.filter(created_at__year=today.year)
        elif self.range_val == 'custom':
            if self.gte_val:
                queryset = queryset.filter(created_at__date__gte=self.gte_val)
            if self.lte_val:
                queryset = queryset.filter(created_at__date__lte=self.lte_val)
        return queryset

    def choices(self, changelist):
        return []



class PostulacionDateRangeFilter(DateRangeFilter):
    title = 'Fecha de postulación'


class ContactoDateRangeFilter(DateRangeFilter):
    title = 'Fecha de contacto'


class EstadoListFilter(admin.SimpleListFilter):
    title = 'Estado'
    parameter_name = 'estado'
    template = 'admin/dropdown_filter.html'

    def lookups(self, request, model_admin):
        return [
            ("Postulando", "Postulando"),
            ("Califica", "Califica"),
            ("No califica", "No califica"),
            ("Contactado", "Contactado"),
            ("Volver a llamar", "Volver a llamar"),
            ("Seguna etapa", "Seguna etapa"),
            ("Contratado", "Contratado"),
            ("Desvinculado", "Desvinculado"),
            ("Abandono", "Abandono"),
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(estado=self.value())
        return queryset


class PuestoListFilter(admin.SimpleListFilter):
    title = 'Puesto al que postula'
    parameter_name = 'puesto'
    template = 'admin/dropdown_filter.html'

    def lookups(self, request, model_admin):
        return [
            ('ejecutivo_encuestas', 'Ejecutivo de Encuestas'),
            ('ejecutivo_ventas', 'Ejecutivo de Ventas de Seguros'),
            ('otros', 'Otros'),
        ]

    def queryset(self, request, queryset):
        val = self.value()
        if val == 'ejecutivo_encuestas':
            return queryset.filter(position='ejecutivo_encuestas')
        elif val == 'ejecutivo_ventas':
            return queryset.filter(position='ejecutivo_ventas')
        elif val == 'otros':
            return queryset.exclude(position__in=['ejecutivo_encuestas', 'ejecutivo_ventas'])
        return queryset


@admin.register(CompanyLead)
class CompanyLeadAdmin(admin.ModelAdmin):
    list_display = (
        "company_name",
        "contact_name",
        "email",
        "phone",
        "service_interest",
        "created_at",
    )
    list_filter = (ContactoDateRangeFilter,)
    search_fields = ("company_name", "contact_name", "email", "phone")
    actions = [export_to_excel]
    list_per_page = 1000000

    class Media:
        css = {
            "all": ("css/admin_custom_v2.css",)
        }
        js = ("js/admin_custom_v2.js",)


@admin.register(Application)
class ApplicationAdmin(admin.ModelAdmin):
    list_display = (
        "full_name",
        "email",
        "phone",
        "estado",
        "observaciones",
        "position",
        "sales_experience",
        "availability",
        "equipamiento_audio",
        "especificaciones_pc",
        "conexion",
        "competencias",
        "cv",
        "created_at",
    )
    list_editable = (
        "estado",
        "observaciones",
    )
    list_filter = (PostulacionDateRangeFilter, EstadoListFilter, PuestoListFilter)
    search_fields = ("full_name", "email", "phone")
    actions = [export_to_excel]
    list_per_page = 1000000

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('ajax-save/', self.admin_site.admin_view(self.ajax_save_view), name='application_ajax_save'),
        ]
        return custom_urls + urls

    def ajax_save_view(self, request):
        from django.http import JsonResponse
        import json
        
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                app_id = data.get('id')
                estado = data.get('estado')
                observaciones = data.get('observaciones')
                
                app = self.model.objects.get(id=app_id)
                if estado is not None:
                    app.estado = estado
                if observaciones is not None:
                    app.observaciones = observaciones
                app.save()
                return JsonResponse({'status': 'success'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        from django.forms import Textarea
        if db_field.name == 'observaciones':
            kwargs['widget'] = Textarea(attrs={'rows': 1, 'style': 'width: 300px; font-size: 0.75rem; resize: vertical;'})
        return super().formfield_for_dbfield(db_field, request, **kwargs)

    class Media:
        css = {
            "all": ("css/admin_custom_v2.css",)
        }
        js = ("js/admin_custom_v2.js",)
